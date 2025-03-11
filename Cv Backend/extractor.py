import convert_pdf_to_images 
import os
from openai import OpenAI
import json
import re
import os
from openpyxl import Workbook, load_workbook

def create_str(data):
    output_str = ''
    
    if isinstance(data, list):  # Checks if the input is a list
        for item in data:
            output_str += f"{item}\n"
    elif isinstance(data, dict):  #Checks if the input is a dictionary
        for k in data:
            output_str += f"{k}: {str(data[k])}\n"
    
    return output_str


def create_or_open_excel(file_path="records.xlsx"):
    """
    Creates a new Excel file with specified columns if it doesn't exist,
    or opens it if it already exists.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        str: Path to the Excel file
    """
    if not os.path.exists(file_path):
        # Creating a new workbook with specified columns
        wb = Workbook()
        ws = wb.active
        ws.title = "Records"
        
        # Defining columns
        columns = [
            "Name", "Email", "Phone Number", 
            "Qualifications",
            "Education", 
            "Projects", 
            "Personal Information"
        ]
        
        # Add headers
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        
        # Save the workbook
        wb.save(file_path)
        print(f"Created new Excel file: {file_path}")
    else:
        print(f"Excel file already exists: {file_path}")
    
    return file_path

def add_record(name, email, phone_number, education, projects, qualifications, personal_info, file_path="./records.xlsx"):
    # Ensures the file exists
    create_or_open_excel(file_path)
    
    # Loading the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Find the next empty row
    next_row = ws.max_row + 1
    
    # Add new record
    ws.cell(row=next_row, column=1, value=name)
    ws.cell(row=next_row, column=2, value=email)
    ws.cell(row=next_row, column=3, value=phone_number)
    ws.cell(row=next_row, column=4, value=qualifications)
    ws.cell(row=next_row, column=5, value=education)
    ws.cell(row=next_row, column=6, value=projects)
    ws.cell(row=next_row, column=7, value=personal_info)
    
    # Save the workbook
    wb.save(file_path)
    print(f"Added new record for {name}")
    return True


openai_api_key = ""

os.environ["OPENAI_API_KEY"] = openai_api_key
client = OpenAI()


def extract_json_dict(text):
    """
    Extract a JSON string delimited by ```json and ``` from a larger text string
    and convert it to a Python dictionary.
    
    Args:
        text (str): The input text containing a JSON string between ```json and ```
        
    Returns:
        dict: The parsed JSON as a Python dictionary
        None: If no valid JSON is found
    """
    # Regular expression to find JSON content between ```json and ```
    pattern = r"```json\s*([\s\S]*?)```"
    
    # Search for the pattern in the text
    match = re.search(pattern, text)
    
    if match:
        json_str = match.group(1).strip()
        try:
            # Parse the JSON string into a Python dictionary
            json_dict = json.loads(json_str)
            return json_dict
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON: {e}")
            return None
    else:
        print("No JSON content found between ```json and ``` markers")
        return None


def extract_informations_from_pdf(file_name:str,name:str,phone:str,email:str):

    imgs =convert_pdf_to_images.convert_pdf_to_images('./uploads/'+file_name,'./images')

    print("images:", len(imgs))


    collected_data=""
    for img in imgs:

        # convert image to base64
        # print(img)

        prompt_text='''
        Based on the given CV Image Extract following informations in a json format:
        If any details missing, leave it as empty.
        ```json{
        "education":"",
        "qualifications":["list of qualifications based on the cv and summarize them here"],
        "projects":[
            {
            "title":"",
            "used_technologies":[],
            "description":""
            }
        ],
        "personal_informations":{
                "name":"",
                "contact_deatils":{
                }
            }
        }```
        '''
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": prompt_text,
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{img}"},
                        },
                    ],
                }
            ],
        )

        content = response.choices[0].message.content
        collected_data+="page content:\n" + content


    summarize_content_prompt= "without prior knowlage combine given page json contents to one json content without changing the format. input content: "
    summarize_content_prompt += collected_data
    response = client.chat.completions.create(
        messages=[{
            "role": "system",
            "content":summarize_content_prompt
        }],
        model="gpt-4o-mini",
    )


    final_data_content_str = response.choices[0].message.content
    print("final_data_content_str:",final_data_content_str)
    output_data_obj = extract_json_dict(final_data_content_str)
    print(output_data_obj)
    education =create_str(output_data_obj['education'])
    projects = create_str(output_data_obj['projects'])
    qualifications=create_str(output_data_obj['qualifications'])
    personal_info=create_str(output_data_obj['personal_informations'])
    add_record(name,
               email,
               phone,
               education,
               projects,
               qualifications,
               personal_info,
               )

